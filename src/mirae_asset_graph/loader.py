from __future__ import annotations

import json
from collections import Counter
from collections.abc import Iterator, Mapping
from datetime import date
from pathlib import Path
from typing import Any

from neo4j import Driver
from neo4j.exceptions import Neo4jError
from openpyxl import load_workbook

from .model import (
    BOND,
    DATASETS,
    DEBT_INSTRUMENT,
    ETF,
    ETN,
    GLOBAL_ETF,
    GLOBAL_ETN,
    KOREAN_BOND,
    KOREAN_ETF,
    KOREAN_ETN,
    LISTED_SECURITY,
    LISTING,
    MUTUAL_FUND,
    NON_TRADABLE_FUND_UNIT,
    PUBLIC_FUND,
    PUBLIC_FUND_UNIT,
    TRADABLE_FUND_UNIT,
    DatasetSpec,
    benchmark,
    boolean,
    class_refs,
    classification,
    clean,
    compact,
    compact_list,
    file_sha256,
    fund_uri,
    identifier,
    is_isin,
    is_korean_source_item,
    number,
    organization,
    parsed_date,
    raw_properties,
    security_uri,
    snapshot_date,
    source_record_uri,
    text,
)

RESOURCE_CONSTRAINT = """
CREATE CONSTRAINT resource_uri IF NOT EXISTS
FOR (resource:Resource) REQUIRE resource.uri IS UNIQUE
"""

SCHEMA_STATEMENTS = (
    """
    CREATE INDEX source_record_location IF NOT EXISTS
    FOR (record:SourceRecord) ON (record.dataset, record.rowNumber)
    """,
    """
    CREATE INDEX observation_as_of IF NOT EXISTS
    FOR (observation:Observation) ON (observation.asOf)
    """,
    """
    CREATE INDEX identifier_value IF NOT EXISTS
    FOR (identifier:Identifier) ON (identifier.value)
    """,
    """
    CREATE FULLTEXT INDEX financial_entity_search IF NOT EXISTS
    FOR (entity:Bond|Fund|Security|Organization|Benchmark)
    ON EACH [entity.name, entity.shortName, entity.englishName, entity.ticker]
    """,
)

ONTOLOGY_MODULES = (
    "common.ttl",
    "bond_kr.ttl",
    "etf_kr.ttl",
    "etf_gl.ttl",
    "fund_pub.ttl",
)

UPSERT_SOURCE_FILE = """
MERGE (dataset:Resource:SourceDataset {uri: $datasetUri})
SET dataset.code = $dataset,
    dataset.name = $datasetName
MERGE (file:Resource:SourceFile {uri: $fileUri})
SET file.name = $filename,
    file.path = $path,
    file.snapshotDate = $snapshotDate,
    file.sha256 = $sha256,
    file.loadedAt = datetime()
MERGE (file)-[:EXTRACT_OF]->(dataset)
"""

UPDATE_SOURCE_FILE_COUNT = """
MATCH (file:SourceFile:Resource {uri: $fileUri})
SET file.rowCount = $rowCount,
    file.loadedAt = datetime()
"""

UPSERT_FIELDS = """
UNWIND $fields AS field
MATCH (dataset:SourceDataset:Resource {uri: $datasetUri})
MERGE (definition:Resource:FieldDefinition {uri: field.uri})
SET definition += field.properties
MERGE (dataset)-[:HAS_FIELD]->(definition)
"""

UPSERT_RAW_RECORDS = """
UNWIND $rows AS row
MATCH (file:SourceFile:Resource {uri: $fileUri})
MERGE (record:Resource:SourceRecord {uri: row.uri})
SET record += row.properties,
    record.dataset = $dataset,
    record.rowNumber = row.rowNumber,
    record.loadedAt = datetime()
MERGE (record)-[:IN_FILE]->(file)
"""

UPSERT_QUALITY_ISSUES = """
UNWIND $rows AS row
MATCH (record:SourceRecord:Resource {uri: row.sourceUri})
SET record:RejectedRecord
MERGE (issue:Resource:DataQualityIssue {uri: row.uri})
SET issue.code = row.code,
    issue.message = row.message,
    issue.dataset = row.dataset,
    issue.rowNumber = row.rowNumber,
    issue.detectedAt = datetime()
MERGE (record)-[:HAS_ISSUE]->(issue)
"""

UPSERT_BONDS = """
UNWIND $rows AS row
MATCH (source:SourceRecord:Resource {uri: row.sourceUri})
MERGE (bond:Resource:Entity:FinancialInstrument:Security:Bond {uri: row.uri})
SET bond += row.properties,
    bond.updatedAt = datetime()
MERGE (source)-[:DESCRIBES]->(bond)
FOREACH (class IN row.classes |
  MERGE (type:Resource {uri: class.uri})
  SET type:OntologyClass, type.name = coalesce(type.name, class.name)
  MERGE (bond)-[:INSTANCE_OF]->(type)
)
FOREACH (item IN row.identifiers |
  MERGE (id:Resource:Identifier {uri: item.uri})
  SET id.scheme = item.scheme, id.value = item.value
  MERGE (bond)-[:HAS_IDENTIFIER]->(id)
)
FOREACH (item IN row.organizations |
  MERGE (org:Resource:Organization {uri: item.uri})
  SET org.name = coalesce(item.name, org.name),
      org.code = coalesce(item.code, org.code),
      org.identityScheme = item.identityScheme
  MERGE (bond)-[issued:ISSUED_BY]->(org)
  SET issued.sourceField = item.sourceField
)
FOREACH (item IN row.classifications |
  MERGE (category:Resource:Classification {uri: item.uri})
  SET category.scheme = item.scheme, category.code = item.code, category.name = item.name
  MERGE (bond)-[:CLASSIFIED_AS]->(category)
)
FOREACH (item IN row.listings |
  MERGE (listing:Resource:Listing {uri: item.uri})
  SET listing += item.properties
  MERGE (bond)-[:LISTED_AS]->(listing)
  MERGE (market:Resource:Market {uri: item.marketUri})
  SET market.code = item.marketCode, market.name = item.marketName
  MERGE (listing)-[:ON_MARKET]->(market)
)
FOREACH (item IN row.observations |
  MERGE (observation:Resource:Observation:BondSnapshot {uri: item.uri})
  SET observation += item.properties
  MERGE (bond)-[:HAS_OBSERVATION]->(observation)
)
"""

UPSERT_FUNDS = """
UNWIND $rows AS row
MATCH (source:SourceRecord:Resource {uri: row.sourceUri})
MERGE (fund:Resource:Entity:Fund {uri: row.fundUri})
SET fund += row.fundProperties,
    fund.updatedAt = datetime()
MERGE (unit:Resource:Entity:FinancialInstrument:Security:FundUnit {uri: row.unitUri})
SET unit += row.unitProperties,
    unit.updatedAt = datetime()
MERGE (fund)-[:HAS_UNIT]->(unit)
MERGE (source)-[:DESCRIBES]->(fund)
MERGE (source)-[:DESCRIBES]->(unit)
FOREACH (class IN row.fundClasses |
  MERGE (type:Resource {uri: class.uri})
  SET type:OntologyClass, type.name = coalesce(type.name, class.name)
  MERGE (fund)-[:INSTANCE_OF]->(type)
)
FOREACH (class IN row.unitClasses |
  MERGE (type:Resource {uri: class.uri})
  SET type:OntologyClass, type.name = coalesce(type.name, class.name)
  MERGE (unit)-[:INSTANCE_OF]->(type)
)
FOREACH (item IN row.identifiers |
  MERGE (id:Resource:Identifier {uri: item.uri})
  SET id.scheme = item.scheme, id.value = item.value
  MERGE (unit)-[:HAS_IDENTIFIER]->(id)
)
FOREACH (item IN row.organizations |
  MERGE (org:Resource:Organization {uri: item.uri})
  SET org.name = coalesce(item.name, org.name),
      org.code = coalesce(item.code, org.code),
      org.identityScheme = item.identityScheme
  MERGE (fund)-[managed:MANAGED_BY]->(org)
  SET managed.sourceField = item.sourceField
)
FOREACH (item IN row.benchmarks |
  MERGE (bench:Resource:Benchmark {uri: item.uri})
  SET bench.name = item.name, bench.englishName = item.englishName
  MERGE (fund)-[:TRACKS]->(bench)
)
FOREACH (item IN row.classifications |
  MERGE (category:Resource:Classification {uri: item.uri})
  SET category.scheme = item.scheme, category.code = item.code, category.name = item.name
  MERGE (unit)-[:CLASSIFIED_AS]->(category)
)
FOREACH (item IN row.listings |
  MERGE (listing:Resource:Listing {uri: item.uri})
  SET listing += item.properties
  MERGE (unit)-[:LISTED_AS]->(listing)
  MERGE (market:Resource:Market {uri: item.marketUri})
  SET market.code = item.marketCode, market.name = item.marketName
  MERGE (listing)-[:ON_MARKET]->(market)
  MERGE (observation:Resource:Observation:MarketSnapshot {uri: item.observationUri})
  SET observation += item.observationProperties
  MERGE (listing)-[:HAS_OBSERVATION]->(observation)
)
FOREACH (item IN row.fundObservations |
  MERGE (observation:Resource:Observation:FundSnapshot {uri: item.uri})
  SET observation += item.properties
  MERGE (unit)-[:HAS_OBSERVATION]->(observation)
)
FOREACH (item IN row.offerings |
  MERGE (offering:Resource:Offering {uri: item.uri})
  SET offering += item.properties
  MERGE (offering)-[:OFFERS]->(unit)
)
"""

UPSERT_ETNS = """
UNWIND $rows AS row
MATCH (source:SourceRecord:Resource {uri: row.sourceUri})
MERGE (note:Resource:Entity:FinancialInstrument:Security:ExchangeTradedNote {uri: row.uri})
SET note += row.properties,
    note.updatedAt = datetime()
MERGE (source)-[:DESCRIBES]->(note)
FOREACH (class IN row.classes |
  MERGE (type:Resource {uri: class.uri})
  SET type:OntologyClass, type.name = coalesce(type.name, class.name)
  MERGE (note)-[:INSTANCE_OF]->(type)
)
FOREACH (item IN row.identifiers |
  MERGE (id:Resource:Identifier {uri: item.uri})
  SET id.scheme = item.scheme, id.value = item.value
  MERGE (note)-[:HAS_IDENTIFIER]->(id)
)
FOREACH (item IN row.organizations |
  MERGE (org:Resource:Organization {uri: item.uri})
  SET org.name = coalesce(item.name, org.name),
      org.code = coalesce(item.code, org.code),
      org.identityScheme = item.identityScheme
  MERGE (note)-[issued:ISSUED_BY]->(org)
  SET issued.sourceField = item.sourceField
)
FOREACH (item IN row.benchmarks |
  MERGE (bench:Resource:Benchmark {uri: item.uri})
  SET bench.name = item.name, bench.englishName = item.englishName
  MERGE (note)-[:TRACKS]->(bench)
)
FOREACH (item IN row.classifications |
  MERGE (category:Resource:Classification {uri: item.uri})
  SET category.scheme = item.scheme, category.code = item.code, category.name = item.name
  MERGE (note)-[:CLASSIFIED_AS]->(category)
)
FOREACH (item IN row.listings |
  MERGE (listing:Resource:Listing {uri: item.uri})
  SET listing += item.properties
  MERGE (note)-[:LISTED_AS]->(listing)
  MERGE (market:Resource:Market {uri: item.marketUri})
  SET market.code = item.marketCode, market.name = item.marketName
  MERGE (listing)-[:ON_MARKET]->(market)
  MERGE (observation:Resource:Observation:MarketSnapshot {uri: item.observationUri})
  SET observation += item.observationProperties
  MERGE (listing)-[:HAS_OBSERVATION]->(observation)
)
FOREACH (item IN row.offerings |
  MERGE (offering:Resource:Offering {uri: item.uri})
  SET offering += item.properties
  MERGE (offering)-[:OFFERS]->(note)
)
"""


def _read_rows(path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(iterator)]
        for row_number, values in enumerate(iterator, 2):
            yield row_number, {key: clean(value) for key, value in zip(headers, values)}
    finally:
        workbook.close()


def _batch(iterator: Iterator[tuple[int, dict[str, Any]]], size: int) -> Iterator[list[tuple[int, dict[str, Any]]]]:
    current: list[tuple[int, dict[str, Any]]] = []
    for item in iterator:
        current.append(item)
        if len(current) >= size:
            yield current
            current = []
    if current:
        yield current


def _numeric_properties(row: Mapping[str, Any], mapping: Mapping[str, str]) -> dict[str, float]:
    return compact({target: number(row.get(source)) for source, target in mapping.items()})


def _market_payload(
    row: Mapping[str, Any],
    *,
    dataset: str,
    unit_uri: str,
    file_snapshot: date | None,
) -> dict[str, Any] | None:
    ticker = text(row.get("pd_itm_no_ma")) or text(row.get("pd_itm_no"))
    market_code = text(row.get("pd_exg_mkt_cd")) or text(row.get("pd_mkt_id"))
    if not ticker or not market_code:
        return None
    market_name = text(row.get("pd_mkt_nm")) or text(row.get("pd_mkt_id")) or market_code
    listing_uri = f"urn:miraeasset:listing:{dataset}:{market_code}:{ticker}"
    as_of = parsed_date(row.get("du_clpr_base_dt")) or parsed_date(row.get("du_upt_dt")) or file_snapshot
    observation_properties = _numeric_properties(
        row,
        {
            "du_bpr": "basePrice",
            "du_opr": "openPrice",
            "du_hpr": "highPrice",
            "du_lpr": "lowPrice",
            "du_clpr": "closePrice",
            "du_diff_rt": "priceChangeRate",
            "du_val_1d": "tradingValue1d",
            "du_vol_1d": "tradingVolume1d",
            "du_vol_avg_5d": "averageVolume5d",
            "du_vol_avg_1m": "averageVolume1m",
            "ru_mkt_price": "realtimeMarketPrice",
            "ru_mkt_volume": "realtimeMarketVolume",
            "nru_mkt_inav": "realtimeIndicativeNav",
            "nru_mkt_diff_rt": "realtimeMarketDifferenceRate",
        },
    )
    observation_properties.update(
        compact(
            {
                "asOf": as_of,
                "sourceDate": parsed_date(row.get("du_upt_dt")),
                "closePriceSource": text(row.get("du_clpr_src")),
                "currency": text(row.get("pd_trd_ccy")) or text(row.get("pd_curr_cd")),
                "dataset": dataset,
            }
        )
    )
    return {
        "uri": listing_uri,
        "properties": compact(
            {
                "ticker": ticker,
                "marketCode": market_code,
                "marketName": market_name,
                "listingDate": parsed_date(row.get("pd_lstg_dt")),
                "listingPrice": number(row.get("pd_lst_price")),
                "listedShareCount": number(row.get("pd_lst_stk_cnt")),
                "dataset": dataset,
                "fiboClassUri": LISTING,
            }
        ),
        "marketUri": f"urn:miraeasset:market:{dataset}:{market_code}",
        "marketCode": market_code,
        "marketName": market_name,
        "observationUri": f"urn:miraeasset:observation:market:{dataset}:{market_code}:{ticker}:{as_of or 'undated'}",
        "observationProperties": observation_properties,
    }


def _fund_observation(
    row: Mapping[str, Any],
    *,
    dataset: str,
    unit_uri: str,
    file_snapshot: date | None,
) -> dict[str, Any]:
    as_of = parsed_date(row.get("du_upt_dt")) or parsed_date(row.get("cu_upt_dt")) or file_snapshot
    properties = _numeric_properties(
        row,
        {
            "du_last_aum": "assetsUnderManagement",
            "du_last_nav": "netAssetValue",
            "du_nav_yday": "previousNav",
            "du_nav_rnf_amt": "navChangeAmount",
            "du_er_1d": "return1d",
            "du_er_5d": "return5d",
            "du_er_1m": "return1m",
            "du_er_3m": "return3m",
            "du_er_6m": "return6m",
            "du_er_1y": "return1y",
            "du_er_ytd": "returnYtd",
            "pd_net_tamt": "netAssetTotal",
            "pd_nav_pshr": "navPerShare",
            "pd_net_ast_pshr": "netAssetsPerShare",
            "pd_net_prft_pshr": "netProfitPerShare",
            "pd_net_rt_ast_pshr": "netReturnAssetsPerShare",
        },
    )
    properties.update(compact({"asOf": as_of, "dataset": dataset}))
    return {
        "uri": f"urn:miraeasset:observation:fund:{dataset}:{unit_uri.rsplit(':', 1)[-1]}:{as_of or 'undated'}",
        "properties": properties,
    }


def _transform_bond(
    row_number: int,
    row: Mapping[str, Any],
    *,
    dataset: str,
    file_snapshot: date | None,
) -> tuple[str, dict[str, Any]]:
    product_number = text(row.get("PD_NO"))
    source_uri = source_record_uri(dataset, file_snapshot, row_number)
    if not product_number:
        return "issue", {
            "sourceUri": source_uri,
            "uri": source_uri + ":issue:missing-product-number",
            "code": "MISSING_PRODUCT_NUMBER",
            "message": "PD_NO is empty; no canonical bond was created.",
            "dataset": dataset,
            "rowNumber": row_number,
        }

    uri = security_uri(product_number, f"bond:{product_number}")
    as_of = parsed_date(row.get("PD_STD_INFO_UPDATE")) or file_snapshot
    observation_properties = _numeric_properties(
        row,
        {
            "BUY_YIELD": "buyYield",
            "CORP_PRETAX_YIELD": "corporatePretaxYield",
            "CORP_AFTER_TAX_YIELD": "corporateAfterTaxYield",
            "AFTER_TAX_YIELD": "afterTaxYield",
            "PREF_TAX_YIELD": "preferentialTaxYield",
            "AVG_ANNUAL_TAX_YIELD": "averageAnnualTaxYield",
            "DEPO_EQUIV_YIELD_154": "depositEquivalentYield154",
            "BUYABLE_QUANTITY": "buyableQuantity",
            "REMAINING_DAYS": "remainingDays",
            "DUR": "durationRaw",
            "COV": "convexityRaw",
            "NDY_DUR": "nextDayDurationRaw",
            "NDY_COV": "nextDayConvexityRaw",
            "EVAL_PRICE": "evaluationPrice",
            "APPLIED_YIELD": "appliedYield",
            "DIRTY": "dirtyPrice",
            "NDY_EVAL_PRICE": "nextDayEvaluationPrice",
            "NDY_APPLIED_YIELD": "nextDayAppliedYield",
            "NDY_DIRTY": "nextDayDirtyPrice",
        },
    )
    observation_properties.update(
        compact(
            {
                "asOf": as_of,
                "creditGrade": text(row.get("CRD_GRD")),
                "creditGradeDate": parsed_date(row.get("CRD_GRD_DT")),
                "dataset": dataset,
                "semanticStatus": "source metric codes retained where definitions are ambiguous",
            }
        )
    )
    identifiers = compact_list(
        [identifier("ISIN" if is_isin(product_number) else "PD_NO", product_number)]
    )
    issuer = organization(row.get("PD_PBCM"), source_field="PD_PBCM")
    classifications = compact_list(
        [
            classification("bond-main-category", row.get("STD_PD_MCLS_NM")),
            classification("bond-subcategory", row.get("STD_PD_SCLS_NM")),
            classification("bond-kind", row.get("BD_KND")),
            classification("credit-grade", row.get("CRD_GRD") or row.get("PD_EVCO_CRD_GRD")),
            classification("risk-code", row.get("PD_RISK_GCD")),
        ]
    )
    listings: list[dict[str, Any]] = []
    if text(row.get("PD_EXG_MKT")) == "장내":
        listings.append(
            {
                "uri": f"urn:miraeasset:listing:{dataset}:krx-bond:{product_number}",
                "properties": {
                    "marketCode": "KRX_BOND",
                    "marketName": "Korean exchange-traded bond market",
                    "dataset": dataset,
                    "fiboClassUri": LISTING,
                },
                "marketUri": "urn:miraeasset:market:kr:krx-bond",
                "marketCode": "KRX_BOND",
                "marketName": "Korean exchange-traded bond market",
            }
        )

    payload = {
        "sourceUri": source_uri,
        "uri": uri,
        "properties": compact(
            {
                "sourceProductNumber": product_number,
                "name": text(row.get("PD_NM")),
                "shortName": text(row.get("PD_ABRV_NM")),
                "englishName": text(row.get("PD_ENG_NM")),
                "shortEnglishName": text(row.get("PD_ABRV_ENG_NM")),
                "countryCode": text(row.get("PD_CTRY_CD")),
                "currency": text(row.get("CURR_CD")),
                "issueOutstandingAmount": number(row.get("ISU_BAL_AMT")),
                "issueDate": parsed_date(row.get("ISU_DT")),
                "maturityDate": parsed_date(row.get("MAT_DT")),
                "couponRate": number(row.get("SRFC_IRT")),
                "exchangeMarket": text(row.get("PD_EXG_MKT")),
                "issuerDisplayName": text(row.get("PD_PBCM")),
                "fiboClassUri": BOND,
                "dataset": dataset,
            }
        ),
        "classes": class_refs([BOND, KOREAN_BOND]),
        "identifiers": identifiers,
        "organizations": compact_list([issuer]),
        "classifications": classifications,
        "listings": listings,
        "observations": [
            {
                "uri": f"urn:miraeasset:observation:bond:{product_number}:{as_of or 'undated'}",
                "properties": observation_properties,
            }
        ],
    }
    return "bond", payload


def _transform_exchange_product(
    row_number: int,
    row: Mapping[str, Any],
    *,
    dataset: str,
    file_snapshot: date | None,
) -> tuple[str, dict[str, Any]]:
    group = (text(row.get("pd_grp_no")) or "").upper()
    local_id = text(row.get("pd_itm_no"))
    isin_value = local_id if dataset == "domestic_etf_etn" else text(row.get("pd_isin_cd"))
    ticker = text(row.get("pd_itm_no_ma")) or local_id
    market_code = text(row.get("pd_exg_mkt_cd")) or "unknown-market"
    source_uri = source_record_uri(dataset, file_snapshot, row_number)
    if not local_id:
        return "issue", {
            "sourceUri": source_uri,
            "uri": source_uri + ":issue:missing-item-number",
            "code": "MISSING_ITEM_NUMBER",
            "message": "pd_itm_no is empty; no canonical exchange product was created.",
            "dataset": dataset,
            "rowNumber": row_number,
        }

    fallback = f"{dataset}:{market_code}:{ticker or local_id}"
    unit_uri = security_uri(isin_value, fallback)
    manager = organization(row.get("cu_fund_mgmt_co"), source_field="cu_fund_mgmt_co")
    base_index = benchmark(row.get("cu_base_index"))
    listing = _market_payload(row, dataset=dataset, unit_uri=unit_uri, file_snapshot=file_snapshot)
    identifiers = compact_list(
        [
            identifier("ISIN", isin_value) if is_isin(isin_value) else None,
            identifier("SOURCE_ITEM", local_id),
            identifier("LISTING_TICKER", ticker),
        ]
    )
    classifications = compact_list(
        [
            classification("product-group", group),
            classification("investment-asset-type", row.get("wu_inv_ast_type")),
            classification("investment-region", row.get("wu_inv_rgn")),
            classification("sector", row.get("pd_sect_cd"), row.get("pd_sect_nm")),
            classification("risk", row.get("pd_risk_cd"), row.get("pd_risk_nm")),
        ]
    )
    offering = {
        "uri": f"urn:miraeasset:offering:{dataset}:{market_code}:{ticker or local_id}",
        "properties": compact(
            {
                "availableForSale": boolean(row.get("pd_sale_yn")),
                "tradingHalted": boolean(row.get("pd_tr_yn")),
                "dataset": dataset,
                "asOf": file_snapshot,
            }
        ),
    }
    common_properties = compact(
        {
            "sourceItemNumber": local_id,
            "name": text(row.get("pd_nm")),
            "shortName": text(row.get("pd_abrv_nm")),
            "currency": text(row.get("pd_curr_cd")) or text(row.get("pd_trd_ccy")),
            "productGroup": group,
            "strategy": text(row.get("cu_strtegy")),
            "baseIndexName": text(row.get("cu_base_index")),
            "leverageFactor": number(row.get("cu_lev_fector")),
            "feeRate": number(row.get("cu_charge_rt")),
            "otherFeeRate": number(row.get("cu_charge_etc_rt")),
            "indexReplicationMethod": text(row.get("cu_index_repl_mthd")),
            "indexTracking": boolean(row.get("cu_index_tracking_yn")),
            "inverseOrShort": boolean(row.get("cu_inverse_short_yn")),
            "dataset": dataset,
        }
    )
    if dataset == "domestic_etf_etn":
        domain_etf_class, domain_etn_class = KOREAN_ETF, KOREAN_ETN
    elif dataset == "overseas_etf_etn":
        domain_etf_class, domain_etn_class = GLOBAL_ETF, GLOBAL_ETN
    else:
        raise ValueError(f"Unsupported exchange-product dataset: {dataset}")

    if group == "ETN" or boolean(row.get("cu_etn_yn")) is True:
        payload = {
            "sourceUri": source_uri,
            "uri": unit_uri,
            "properties": {**common_properties, "fiboClassUri": ETN},
            "classes": class_refs(
                [ETN, domain_etn_class, DEBT_INSTRUMENT, LISTED_SECURITY]
            ),
            "identifiers": identifiers,
            "organizations": compact_list([manager]),
            "benchmarks": compact_list([base_index]),
            "classifications": classifications,
            "listings": compact_list([listing]),
            "offerings": [offering],
        }
        return "etn", payload

    fund_id = fund_uri(isin_value, fallback)
    payload = {
        "sourceUri": source_uri,
        "fundUri": fund_id,
        "unitUri": unit_uri,
        "fundProperties": {**common_properties, "fiboClassUri": ETF},
        "unitProperties": compact(
            {
                "name": text(row.get("pd_nm")),
                "shortName": text(row.get("pd_abrv_nm")),
                "currency": text(row.get("pd_curr_cd")) or text(row.get("pd_trd_ccy")),
                "sourceItemNumber": local_id,
                "fiboClassUri": TRADABLE_FUND_UNIT,
                "dataset": dataset,
            }
        ),
        "fundClasses": class_refs([ETF, domain_etf_class]),
        "unitClasses": class_refs([TRADABLE_FUND_UNIT, LISTED_SECURITY]),
        "identifiers": identifiers,
        "organizations": compact_list([manager]),
        "benchmarks": compact_list([base_index]),
        "classifications": classifications,
        "listings": compact_list([listing]),
        "fundObservations": [
            _fund_observation(row, dataset=dataset, unit_uri=unit_uri, file_snapshot=file_snapshot)
        ],
        "offerings": [offering],
    }
    return "fund", payload


def _transform_public_fund(
    row_number: int,
    row: Mapping[str, Any],
    *,
    dataset: str,
    file_snapshot: date | None,
) -> tuple[str, dict[str, Any]]:
    item_number = text(row.get("itm_no"))
    source_uri = source_record_uri(dataset, file_snapshot, row_number)
    if not is_korean_source_item(item_number):
        return "issue", {
            "sourceUri": source_uri,
            "uri": source_uri + ":issue:invalid-item-number",
            "code": "INVALID_ITEM_NUMBER",
            "message": f"itm_no is not an ISIN-shaped identifier: {item_number!r}; no canonical fund unit was created.",
            "dataset": dataset,
            "rowNumber": row_number,
        }

    name = text(row.get("itm_nm")) or item_number
    short_name = text(row.get("itm_abrv_nm"))
    is_etf = "상장지수" in name or "ETF" in name.upper() or bool(short_name and "ETF" in short_name.upper())
    # itm_no is shaped like an ISIN, but the source documentation does not
    # establish it as an ISIN. Keep it in a source-specific identity domain.
    unit_uri = security_uri(None, f"public-fund:{item_number}")
    fund_id = fund_uri(None, f"public-fund:{item_number}")
    manager = organization(row.get("or_co_xtn_itt_cd"), scheme="source-code", source_field="or_co_xtn_itt_cd")
    base_index = benchmark(row.get("bmrk_nm"), row.get("bmrk_eng_nm"))
    identifiers = compact_list(
        [
            identifier("SOURCE_ITEM", item_number),
            identifier("STD_ITEM", row.get("std_itm_no")),
            identifier("KSD_ITEM", row.get("ksd_itm_no")),
            identifier("REPRESENTATIVE_KSD_ITEM", row.get("rptt_ksd_itm_no")),
            identifier("FSS_ITEM", row.get("fss_itm_no")),
            identifier("MIRAE_ITEM", row.get("mtco_itm_no")),
        ]
    )
    classifications = compact_list(
        [
            classification("product-attribute", row.get("prfd_attr_cd")),
            classification("kofia-fund-class", row.get("kofia_fd_ccd")),
            classification("fund-set-product-code", row.get("fd_set_pcd")),
            classification("investment-region", row.get("fd_ivst_rgn_desc")),
            classification("fund-type", row.get("or_attr_desc")),
            classification("investor-kind", row.get("pers_corp_desc")),
            classification("risk-grade", row.get("zrin_fd_ivst_risk_gcd"), row.get("zrin_fd_ivst_risk_grd_nm")),
        ]
    )
    as_of = file_snapshot
    fund_observation = {
        "uri": f"urn:miraeasset:observation:fund:{dataset}:{item_number}:{as_of or 'undated'}",
        "properties": {
            **_numeric_properties(
                row,
                {
                    "fd_nast_suma": "netAssetTotal",
                    "fd_wk1_ern_r": "return1w",
                    "fd_mm1_ern_r": "return1m",
                    "fd_mm3_ern_r": "return3m",
                    "fd_mm6_ern_r": "return6m",
                    "fd_mm18_ern_r": "return18m",
                    "fd_yr1_ern_r": "return1y",
                    "fd_yr2_ern_r": "return2y",
                    "fd_yr3_ern_r": "return3y",
                    "fd_yr5_ern_r": "return5y",
                },
            ),
            **compact({"asOf": as_of, "dataset": dataset}),
        },
    }
    listing: dict[str, Any] | None = None
    if is_etf:
        listing = {
            "uri": f"urn:miraeasset:listing:{dataset}:krx-unspecified:{item_number}",
            "properties": {
                "marketCode": "KRX_UNSPECIFIED",
                "marketName": "Korea Exchange (venue not supplied)",
                "dataset": dataset,
                "listingInferredFromName": True,
                "fiboClassUri": LISTING,
            },
            "marketUri": "urn:miraeasset:market:kr:krx-unspecified",
            "marketCode": "KRX_UNSPECIFIED",
            "marketName": "Korea Exchange (venue not supplied)",
            "observationUri": f"urn:miraeasset:observation:market:{dataset}:{item_number}:{as_of or 'undated'}",
            "observationProperties": compact({"asOf": as_of, "dataset": dataset}),
        }

    payload = {
        "sourceUri": source_uri,
        "fundUri": fund_id,
        "unitUri": unit_uri,
        "fundProperties": compact(
            {
                "name": name,
                "shortName": short_name,
                "englishName": text(row.get("itm_eng_nm")),
                "currency": text(row.get("curr_cd")),
                "countryCode": text(row.get("fd_estb_ctry_cd")),
                "currencyHedged": boolean(row.get("exchdg_yn")),
                "overseasFund": boolean(row.get("ofsfd_yn")),
                "privateFundDescription": text(row.get("prvo_fd_desc")),
                "publicPrivateDescription": text(row.get("prvo_pbff_desc")),
                "exchangeTradedClassification": "name heuristic" if is_etf else None,
                "fiboClassUri": ETF if is_etf else MUTUAL_FUND,
                "dataset": dataset,
            }
        ),
        "unitProperties": compact(
            {
                "sourceItemNumber": item_number,
                "name": name,
                "shortName": short_name,
                "englishName": text(row.get("itm_eng_nm")),
                "currency": text(row.get("curr_cd")),
                "representativeKsdItemNumber": text(row.get("rptt_ksd_itm_no")),
                "fiboClassUri": TRADABLE_FUND_UNIT if is_etf else NON_TRADABLE_FUND_UNIT,
                "dataset": dataset,
            }
        ),
        "fundClasses": class_refs([ETF if is_etf else MUTUAL_FUND, PUBLIC_FUND]),
        "unitClasses": class_refs(
            [TRADABLE_FUND_UNIT, LISTED_SECURITY, PUBLIC_FUND_UNIT]
            if is_etf
            else [NON_TRADABLE_FUND_UNIT, PUBLIC_FUND_UNIT]
        ),
        "identifiers": identifiers,
        "organizations": compact_list([manager]),
        "benchmarks": compact_list([base_index]),
        "classifications": classifications,
        "listings": compact_list([listing]),
        "fundObservations": [fund_observation],
        "offerings": [
            {
                "uri": f"urn:miraeasset:offering:{dataset}:{item_number}",
                "properties": compact(
                    {
                        "saleStatus": text(row.get("sale_yn")),
                        "availableForSale": boolean(row.get("sale_yn")),
                        "availableThroughFirm": boolean(row.get("thco_sale_yn")),
                        "saleControlType": text(row.get("pfiv_sale_cntl_tcd")),
                        "dataset": dataset,
                        "asOf": as_of,
                    }
                ),
            }
        ],
    }
    return "fund", payload


TRANSFORMERS = {
    "domestic_bonds": _transform_bond,
    "domestic_etf_etn": _transform_exchange_product,
    "overseas_etf_etn": _transform_exchange_product,
    "public_funds": _transform_public_fund,
}


class FinancialProductsLoader:
    def __init__(
        self,
        driver: Driver,
        *,
        input_dir: Path,
        ontology_path: Path,
        batch_size: int = 500,
        database: str = "neo4j",
    ) -> None:
        self.driver = driver
        self.input_dir = input_dir
        self.ontology_path = ontology_path
        self.batch_size = batch_size
        self.database = database
        self.stats: Counter[str] = Counter()

    def prepare(self) -> None:
        self.driver.execute_query(RESOURCE_CONSTRAINT, database_=self.database)
        self._ensure_n10s_graph_config()
        self._import_application_profile()
        for statement in SCHEMA_STATEMENTS:
            self.driver.execute_query(statement, database_=self.database)

    def _ensure_n10s_graph_config(self) -> None:
        try:
            records, _, _ = self.driver.execute_query(
                "CALL n10s.graphconfig.show() YIELD param, value RETURN param, value",
                database_=self.database,
            )
        except Neo4jError:
            records = []
        if records:
            return
        self.driver.execute_query(
            """
            CALL n10s.graphconfig.init({
              handleVocabUris: 'KEEP',
              handleMultival: 'ARRAY',
              keepLangTag: true,
              keepCustomDataTypes: true
            })
            """,
            database_=self.database,
        )

    def _import_application_profile(self) -> None:
        for path in self._ontology_files():
            ttl = path.read_text(encoding="utf-8")
            records, _, _ = self.driver.execute_query(
                """
                CALL n10s.onto.import.inline($ttl, 'Turtle', {
                  classLabel: 'Class',
                  subClassOfRel: 'SUBCLASS_OF',
                  objectPropertyLabel: 'Relationship',
                  dataTypePropertyLabel: 'Property',
                  subPropertyOfRel: 'SUBPROPERTY_OF',
                  domainRel: 'DOMAIN',
                  rangeRel: 'RANGE',
                  addResourceLabels: true
                })
                YIELD terminationStatus, triplesLoaded, triplesParsed
                RETURN terminationStatus, triplesLoaded, triplesParsed
                """,
                ttl=ttl,
                database_=self.database,
            )
            self.stats["ontologyFilesImported"] += 1
            if records:
                triples_loaded = int(records[0]["triplesLoaded"])
                self.stats["ontologyTriplesLoaded"] += triples_loaded
                self.stats[f"ontology.{path.stem}.triplesLoaded"] += triples_loaded
        # With KEEP URI handling, n10s expands its physical schema label and
        # relationship names. Add stable application-level aliases for Cypher
        # and GraphRAG without removing the lossless n10s representation.
        self.driver.execute_query(
            """
            MATCH (type:Resource)
            WHERE 'neo4j://graph.schema#Class' IN labels(type)
            SET type:OntologyClass,
                type.name = coalesce(type.name, type.`neo4j://graph.schema#name`)
            """,
            database_=self.database,
        )
        self.driver.execute_query(
            """
            MATCH (child:Resource)-[:`neo4j://graph.schema#SCO`]->(parent:Resource)
            MERGE (child)-[:SUBCLASS_OF]->(parent)
            """,
            database_=self.database,
        )

    def _ontology_files(self) -> list[Path]:
        if self.ontology_path.is_file():
            return [self.ontology_path]
        if not self.ontology_path.is_dir():
            raise FileNotFoundError(
                f"Ontology path does not exist: {self.ontology_path}"
            )

        paths = [self.ontology_path / name for name in ONTOLOGY_MODULES]
        missing = [path.name for path in paths if not path.is_file()]
        if missing:
            missing_names = ", ".join(missing)
            raise FileNotFoundError(
                f"Ontology directory {self.ontology_path} is missing required modules: "
                f"{missing_names}"
            )
        return paths

    def load(self, selected: set[str] | None = None) -> dict[str, Any]:
        for spec in DATASETS:
            if selected and spec.code not in selected:
                continue
            self._load_dataset(spec)
        return dict(self.stats)

    def dry_run(self, selected: set[str] | None = None) -> dict[str, Any]:
        for spec in DATASETS:
            if selected and spec.code not in selected:
                continue
            path = self._one_file(spec.data_pattern)
            stamp = snapshot_date(path)
            counts: Counter[str] = Counter()
            transformer = TRANSFORMERS[spec.code]
            for row_number, row in _read_rows(path):
                kind, _ = transformer(
                    row_number,
                    row,
                    dataset=spec.code,
                    file_snapshot=stamp,
                )
                counts[kind] += 1
            for kind, count in counts.items():
                self.stats[f"{spec.code}.{kind}"] = count
            self.stats[f"{spec.code}.rawRows"] = sum(counts.values())
            print(f"{spec.code}: {dict(counts)}")
        return dict(self.stats)

    def _load_dataset(self, spec: DatasetSpec) -> None:
        data_path = self._one_file(spec.data_pattern)
        schema_path = self._one_file(spec.schema_pattern)
        stamp = snapshot_date(data_path)
        file_uri = f"urn:miraeasset:source-file:{spec.code}:{stamp or 'undated'}:{data_path.name}"
        dataset_uri = f"urn:miraeasset:dataset:{spec.code}"
        self.driver.execute_query(
            UPSERT_SOURCE_FILE,
            datasetUri=dataset_uri,
            dataset=spec.code,
            datasetName=spec.name,
            fileUri=file_uri,
            filename=data_path.name,
            path=str(data_path.resolve()),
            snapshotDate=stamp,
            sha256=file_sha256(data_path),
            database_=self.database,
        )
        self._load_schema_fields(dataset_uri, spec.code, schema_path)

        transformer = TRANSFORMERS[spec.code]
        count = 0
        for rows in _batch(_read_rows(data_path), self.batch_size):
            raw_batch: list[dict[str, Any]] = []
            canonical: dict[str, list[dict[str, Any]]] = {
                "bond": [],
                "fund": [],
                "etn": [],
                "issue": [],
            }
            for row_number, row in rows:
                source_uri = source_record_uri(spec.code, stamp, row_number)
                raw_batch.append(
                    {
                        "uri": source_uri,
                        "rowNumber": row_number,
                        "properties": raw_properties(row),
                    }
                )
                kind, payload = transformer(
                    row_number,
                    row,
                    dataset=spec.code,
                    file_snapshot=stamp,
                )
                canonical[kind].append(payload)
                self.stats[f"{spec.code}.{kind}"] += 1

            self.driver.execute_query(
                UPSERT_RAW_RECORDS,
                rows=raw_batch,
                fileUri=file_uri,
                dataset=spec.code,
                database_=self.database,
            )
            if canonical["bond"]:
                self.driver.execute_query(UPSERT_BONDS, rows=canonical["bond"], database_=self.database)
            if canonical["fund"]:
                self.driver.execute_query(UPSERT_FUNDS, rows=canonical["fund"], database_=self.database)
            if canonical["etn"]:
                self.driver.execute_query(UPSERT_ETNS, rows=canonical["etn"], database_=self.database)
            if canonical["issue"]:
                self.driver.execute_query(
                    UPSERT_QUALITY_ISSUES,
                    rows=canonical["issue"],
                    database_=self.database,
                )
            count += len(rows)
            self.stats[f"{spec.code}.rawRows"] = count
            if count % (self.batch_size * 10) == 0:
                print(f"{spec.code}: loaded {count:,} source rows", flush=True)

        self.driver.execute_query(
            UPDATE_SOURCE_FILE_COUNT,
            fileUri=file_uri,
            rowCount=count,
            database_=self.database,
        )
        print(f"{spec.code}: complete ({count:,} source rows)", flush=True)

    def _load_schema_fields(self, dataset_uri: str, dataset: str, path: Path) -> None:
        workbook = load_workbook(path, read_only=True, data_only=True)
        fields: list[dict[str, Any]] = []
        try:
            worksheet = workbook.worksheets[0]
            header_row = None
            for row_number, values in enumerate(worksheet.iter_rows(values_only=True), 1):
                if text(values[0] if values else None) == "컬럼명":
                    header_row = row_number
                    break
            if header_row is None:
                raise ValueError(f"Could not find schema header in {path}")
            for ordinal, values in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True), 1
            ):
                field_name = text(values[0] if values else None)
                if not field_name:
                    continue
                fields.append(
                    {
                        "uri": f"urn:miraeasset:field:{dataset}:{field_name}",
                        "properties": compact(
                            {
                                "dataset": dataset,
                                "name": field_name,
                                "ordinal": ordinal,
                                "keyRole": text(values[1]) if len(values) > 1 else None,
                                "sourceDataType": text(values[2]) if len(values) > 2 else None,
                                "koreanName": text(values[3]) if len(values) > 3 else None,
                                "example": text(values[4]) if len(values) > 4 else None,
                            }
                        ),
                    }
                )
        finally:
            workbook.close()
        self.driver.execute_query(
            UPSERT_FIELDS,
            fields=fields,
            datasetUri=dataset_uri,
            database_=self.database,
        )
        self.stats[f"{dataset}.fieldDefinitions"] = len(fields)

    def _one_file(self, pattern: str) -> Path:
        matches = sorted(self.input_dir.glob(pattern))
        if len(matches) != 1:
            raise ValueError(f"Expected one file for {pattern!r}, found {len(matches)}: {matches}")
        return matches[0]

    def validate(self) -> dict[str, Any]:
        query = """
        CALL () { MATCH (record:SourceRecord) RETURN count(record) AS sourceRecords }
        CALL () { MATCH (record:RejectedRecord) RETURN count(record) AS rejectedRecords }
        CALL () { MATCH (bond:Bond) RETURN count(bond) AS bonds }
        CALL () { MATCH (fund:Fund) RETURN count(fund) AS funds }
        CALL () { MATCH (unit:FundUnit) RETURN count(unit) AS fundUnits }
        CALL () { MATCH (note:ExchangeTradedNote) RETURN count(note) AS exchangeTradedNotes }
        CALL () { MATCH (listing:Listing) RETURN count(listing) AS listings }
        CALL () { MATCH (observation:Observation) RETURN count(observation) AS observations }
        CALL () {
          MATCH (record:SourceRecord)
          WHERE NOT (record)-[:DESCRIBES]->() AND NOT record:RejectedRecord
          RETURN count(record) AS unlinkedNonRejectedSourceRecords
        }
        CALL () {
          MATCH (resource:Resource)
          WITH resource.uri AS uri, count(*) AS occurrences
          WHERE occurrences > 1
          RETURN count(*) AS duplicateResourceUris
        }
        RETURN sourceRecords, rejectedRecords, bonds, funds, fundUnits,
               exchangeTradedNotes, listings, observations,
               unlinkedNonRejectedSourceRecords, duplicateResourceUris
        """
        records, _, _ = self.driver.execute_query(query, database_=self.database)
        metrics = dict(records[0]) if records else {}

        dataset_query = """
        MATCH (record:SourceRecord)
        RETURN record.dataset AS dataset,
               count(*) AS sourceRecords,
               sum(CASE WHEN EXISTS { MATCH (record)-[:DESCRIBES]->() } THEN 1 ELSE 0 END) AS linkedRecords,
               sum(CASE WHEN EXISTS { MATCH (record)-[:HAS_ISSUE]->() } THEN 1 ELSE 0 END) AS rejectedRecords
        ORDER BY dataset
        """
        dataset_records, _, _ = self.driver.execute_query(dataset_query, database_=self.database)
        return {
            "metrics": metrics,
            "datasets": [dict(record) for record in dataset_records],
        }

    @staticmethod
    def format_report(report: Mapping[str, Any]) -> str:
        return json.dumps(report, ensure_ascii=False, indent=2, default=str)
