"""Korean manager-published basket adapter for the Phase 3 external-evidence pipeline.

KRX basket automation is NEEDS-CONTRACT, so the approved GO fallback recorded
in docs/external/source-decisions-phase3-2026-08-19.md is the holdings file the fund manager
publishes directly on the official PLUS page: a CSV or XLSX document. This
adapter fetches those documents over plain ``urllib.request`` (stdlib only)
with an identifying User-Agent, a conservative default request interval, and
bounded retries, then normalizes rows into HoldingsRecord JSONL with a
deterministic per-row quarantine path.

Only manager-published evidence is represented; nothing here claims a
regulatory filing status, and PDF is out of scope — format_hint accepts csv
and xlsx only. Reported percent weights are divided by 100 and never
renormalized; a weight is only derived from the document's own positive
market values when the row publishes none. Identity resolution goes
exclusively through the reviewed IdentifierResolver (source ISIN first, then
a reviewed krx_code crosswalk); names are provenance and never a merge key,
and a KRX code is never expanded into an ISIN.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import tempfile
import time
import urllib.error
import urllib.request
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from ..model import is_isin
from ..model import file_sha256
from .base import Adapter
from .artifacts import publish_immutable
from .http_fetch import read_with_validated_redirects
from .records import HoldingsRecord, serialize_jsonl, write_jsonl
from .resolver import IdentifierResolver
from .source_policy import normalize_reviewed_hosts, validate_source_url

MANAGER_BASKET_SOURCE = "manager_basket"

FORMAT_HINTS: tuple[str, ...] = ("csv", "xlsx")

# Fair-access defaults recorded in docs/external/source-decisions-phase3-2026-08-19.md.
DEFAULT_REQUEST_INTERVAL_SECONDS = 0.25
DEFAULT_RETRY_COUNT = 3

_RETRIABLE_STATUS_CODES: frozenset[int] = frozenset({429, 500, 502, 503, 504})
_REQUEST_TIMEOUT_SECONDS = 30.0
_BACKOFF_BASE_SECONDS = 1.0
_BACKOFF_MAX_SECONDS = 30.0

_CSV_ACCEPT = "text/csv"
_XLSX_ACCEPT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_KRX_CODE = "krx_code"
_SOURCE_PUBLISHED = "source_published"
_DERIVED_FROM_VALUE = "derived_from_value"
_DEFAULT_CURRENCY = "KRW"

# Canonical column -> published header aliases. Matching normalizes case and
# all whitespace, so " Weight_Pct " and "비중(%)" both resolve.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("종목명", "구성종목명", "name"),
    "source_isin": ("isin", "표준코드"),
    "krx_code": ("종목코드", "단축코드", "security_code"),
    "weight_pct": ("비중", "비중(%)", "weight_pct"),
    "quantity": ("수량", "quantity"),
    "market_value": ("평가금액", "market_value"),
    "currency": ("통화", "currency"),
    "as_of": ("기준일", "as_of"),
}


@dataclass(frozen=True)
class ManagerBasketTarget:
    """One manager-published holdings document to fetch and normalize."""

    manager_code: str
    fund_code: str
    fund_isin: str
    source_url: str
    as_of: date
    published_at: datetime
    format_hint: str
    reviewed_hosts: tuple[str, ...]
    allow_test_hosts: bool = False

    def __post_init__(self) -> None:
        for field_name in ("manager_code", "fund_code", "fund_isin", "source_url"):
            value = getattr(self, field_name)
            if not isinstance(value, str) or not value.strip():
                raise ValueError(f"ManagerBasketTarget: {field_name} must be a nonempty string")
        for field_name in ("manager_code", "fund_code"):
            value = getattr(self, field_name)
            if "/" in value or "\\" in value or value.strip().startswith("."):
                raise ValueError(
                    f"ManagerBasketTarget: {field_name} becomes a raw path segment and "
                    f"must not contain path separators: {value!r}"
                )
        if not is_isin(self.fund_isin):
            raise ValueError(f"ManagerBasketTarget: fund_isin is not a valid ISIN: {self.fund_isin!r}")
        if not self.source_url.strip().lower().startswith("https://"):
            raise ValueError(
                f"ManagerBasketTarget: source_url must be an https URL: {self.source_url!r}"
            )
        if isinstance(self.as_of, datetime) or not isinstance(self.as_of, date):
            raise ValueError("ManagerBasketTarget: as_of must be a date, not a datetime")
        if not isinstance(self.published_at, datetime):
            raise ValueError("ManagerBasketTarget: published_at must be a datetime")
        if self.format_hint not in FORMAT_HINTS:
            raise ValueError(
                f"ManagerBasketTarget: format_hint must be one of {list(FORMAT_HINTS)}: "
                f"{self.format_hint!r}"
            )
        normalized_hosts = normalize_reviewed_hosts(self.reviewed_hosts)
        object.__setattr__(self, "reviewed_hosts", normalized_hosts)
        normalized_url = validate_source_url(
            MANAGER_BASKET_SOURCE,
            self.source_url,
            manager_hosts=normalized_hosts,
            allow_test_hosts=self.allow_test_hosts,
        )
        if normalized_url != self.source_url:
            object.__setattr__(self, "source_url", normalized_url)


@dataclass(frozen=True)
class BasketQuarantineEntry:
    """One rejected basket row, preserved with its reason and source identifier."""

    source_document_id: str
    constituent_name: str | None
    reason: str
    source_identifier_type: str | None
    source_identifier_value: str | None

    def to_dict(self) -> dict[str, object]:
        """JSON-serializable mapping with the canonical field keys, in order."""
        return {
            "source_document_id": self.source_document_id,
            "constituent_name": self.constituent_name,
            "reason": self.reason,
            "source_identifier_type": self.source_identifier_type,
            "source_identifier_value": self.source_identifier_value,
        }


def _validated_user_agent(value: object) -> str:
    """Require a User-Agent that identifies the caller.

    Accepts a nonempty string containing a contact email address (``@``) or a
    contact URL (``http://`` or ``https://``). Generic browser or library
    strings are rejected because they do not identify the requester.
    """
    if not isinstance(value, str) or not value.strip():
        raise ValueError(
            "ManagerBasketAdapter: user_agent must be a nonempty string that identifies the caller"
        )
    text = value.strip()
    lowered = text.lower()
    if "@" not in text and "http://" not in lowered and "https://" not in lowered:
        raise ValueError(
            "ManagerBasketAdapter: user_agent must contain a contact email address or URL "
            f"for source-identification; got: {value!r}"
        )
    return text


def _atomic_write(path: Path, payload: bytes) -> None:
    """Write payload to a temp file in path's directory, then replace path."""
    path.parent.mkdir(parents=True, exist_ok=True)
    file_descriptor, temp_name = tempfile.mkstemp(dir=path.parent, prefix=path.name, suffix=".tmp")
    temp_path = Path(temp_name)
    try:
        with os.fdopen(file_descriptor, "wb") as handle:
            handle.write(payload)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_path, path)
    except BaseException:
        temp_path.unlink(missing_ok=True)
        raise


def _write_quarantine_jsonl(entries: Iterable[BasketQuarantineEntry], path: Path) -> int:
    """Overwrite path with one deterministic quarantine JSON object per line."""
    written = 0
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for entry in entries:
            handle.write(json.dumps(entry.to_dict(), ensure_ascii=False) + "\n")
            written += 1
    return written


def _serialize_quarantine_jsonl(entries: Iterable[BasketQuarantineEntry]) -> tuple[bytes, int]:
    selected = tuple(entries)
    text = "".join(json.dumps(entry.to_dict(), ensure_ascii=False) + "\n" for entry in selected)
    return text.encode("utf-8"), len(selected)


def _normalized_header(text: str) -> str:
    """Case- and whitespace-insensitive header key (internal spaces removed)."""
    return "".join(text.split()).casefold()


_ALIAS_TO_CANONICAL: dict[str, str] = {
    _normalized_header(alias): canonical
    for canonical, aliases in _COLUMN_ALIASES.items()
    for alias in aliases
}


@dataclass(frozen=True)
class _Row:
    """One parsed basket row, preserving raw texts for error messages.

    A numeric field is absent when its ``*_raw`` text is None, present but
    invalid when the raw text is set while the parsed value is None.
    """

    row_number: int
    name: str | None
    source_isin: str | None
    krx_code: str | None
    weight_pct: float | None
    weight_pct_raw: str | None
    quantity: float | None
    quantity_raw: str | None
    market_value: float | None
    market_value_raw: str | None
    currency: str | None
    as_of_raw: str | None


def _parse_number(raw: str | None) -> tuple[str | None, float | None]:
    """Parse optional numeric text into (raw, value), stripping commas and percents."""
    if raw is None or not raw.strip():
        return None, None
    stripped = raw.strip()
    cleaned = stripped.replace(",", "").replace("%", "").strip()
    try:
        return stripped, float(cleaned)
    except ValueError:
        return stripped, None


def _cell_text(value: object) -> str:
    """Stringify one spreadsheet cell: dates ISO-formatted, integral floats trimmed."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_row(cells: dict[str, str], row_number: int) -> _Row:
    """Build one _Row from the canonical-column cell texts of a data row."""
    weight_pct_raw, weight_pct = _parse_number(cells.get("weight_pct"))
    quantity_raw, quantity = _parse_number(cells.get("quantity"))
    market_value_raw, market_value = _parse_number(cells.get("market_value"))
    return _Row(
        row_number=row_number,
        name=cells.get("name") or None,
        source_isin=cells.get("source_isin") or None,
        krx_code=cells.get("krx_code") or None,
        weight_pct=weight_pct,
        weight_pct_raw=weight_pct_raw,
        quantity=quantity,
        quantity_raw=quantity_raw,
        market_value=market_value,
        market_value_raw=market_value_raw,
        currency=cells.get("currency") or None,
        as_of_raw=cells.get("as_of") or None,
    )


def _decode_csv(payload: bytes, document_label: str) -> str:
    """Decode basket CSV bytes as UTF-8 (with optional BOM), falling back to CP949."""
    try:
        return payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass
    try:
        return payload.decode("cp949")
    except UnicodeDecodeError as error:
        raise ValueError(
            f"manager basket document {document_label} is neither UTF-8 nor CP949: {error}"
        ) from error


def _read_csv_rows(raw_path: Path, document_label: str) -> list[list[str]]:
    """Nonblank CSV rows (header first) decoded with the encoding fallback."""
    text = _decode_csv(raw_path.read_bytes(), document_label)
    reader = csv.reader(io.StringIO(text, newline=""))
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError(f"manager basket document {document_label} contains no header row")
    return rows


def _read_xlsx_rows(raw_path: Path) -> list[list[str]]:
    """Nonblank rows of the first worksheet, every cell stringified."""
    workbook = load_workbook(raw_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = [
            [_cell_text(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
            if any(_cell_text(cell) for cell in row)
        ]
    finally:
        workbook.close()
    if not rows:
        raise ValueError("manager basket XLSX first worksheet contains no header row")
    return rows


def _read_rows(target: ManagerBasketTarget, raw_path: Path) -> list[_Row]:
    """Parse the raw document into _Rows via its target-declared format."""
    document_label = f"{target.manager_code}/{target.fund_code}/{target.as_of.isoformat()}"
    if target.format_hint == "csv":
        table = _read_csv_rows(raw_path, document_label)
    else:
        table = _read_xlsx_rows(raw_path)
    header, data_rows = table[0], table[1:]

    mapping: dict[str, int] = {}
    for index, cell in enumerate(header):
        canonical = _ALIAS_TO_CANONICAL.get(_normalized_header(cell))
        if canonical is not None and canonical not in mapping:
            mapping[canonical] = index
    if "name" not in mapping:
        raise ValueError(
            f"manager basket document {document_label} has no recognized constituent-name "
            f"column (aliases: {', '.join(_COLUMN_ALIASES['name'])})"
        )
    if "source_isin" not in mapping and "krx_code" not in mapping:
        raise ValueError(
            f"manager basket document {document_label} has neither a source-ISIN column "
            f"(aliases: {', '.join(_COLUMN_ALIASES['source_isin'])}) nor a KRX security-code "
            f"column (aliases: {', '.join(_COLUMN_ALIASES['krx_code'])})"
        )
    return [
        _parse_row(
            {canonical: (cells[index].strip() if index < len(cells) else "") for canonical, index in mapping.items()},
            row_number,
        )
        for row_number, cells in enumerate(data_rows, start=2)
    ]


def _parse_as_of(raw: str) -> date:
    """Parse a Korean-published as-of date (``-``, ``.``, or ``/`` separated).

    Spreadsheet cells that Excel formatted as dates arrive as ISO datetime
    text with a midnight component, so a full ISO datetime is also accepted
    and truncated to its date.
    """
    normalized = raw.strip().replace(".", "-").replace("/", "-")
    try:
        return date.fromisoformat(normalized)
    except ValueError:
        pass
    try:
        return datetime.fromisoformat(normalized).date()
    except ValueError as error:
        raise ValueError(f"manager basket as-of value is not a valid date: {raw!r}") from error


def _document_total_market_value(rows: list[_Row]) -> float | None:
    """Sum of the document's positive market values; None when there are none.

    A weight may only be derived from values the document itself publishes,
    and only positive values enter the total — reported weights are never
    renormalized against it.
    """
    total = 0.0
    for row in rows:
        if row.market_value is not None and row.market_value > 0:
            total += row.market_value
    return total if total > 0 else None


class ManagerBasketAdapter(Adapter[ManagerBasketTarget, tuple[Path, Path, int, int]]):
    """Fetch and normalize manager-published Korean basket files. Never touches the graph."""

    source = MANAGER_BASKET_SOURCE

    def __init__(
        self,
        *,
        targets: Iterable[ManagerBasketTarget],
        raw_root: Path,
        resolver: IdentifierResolver,
        user_agent: str,
        refresh: bool,
        request_interval_seconds: float = DEFAULT_REQUEST_INTERVAL_SECONDS,
        retry_count: int = DEFAULT_RETRY_COUNT,
        window_start: date | None = None,
        window_end: date | None = None,
        cutoff: datetime | None = None,
    ) -> None:
        self._targets: tuple[ManagerBasketTarget, ...] = tuple(targets)
        self.raw_root = Path(raw_root)
        self._resolver = resolver
        self.user_agent = _validated_user_agent(user_agent)
        self.refresh = refresh
        for target in self._targets:
            _ = validate_source_url(
                MANAGER_BASKET_SOURCE,
                target.source_url,
                manager_hosts=target.reviewed_hosts,
                allow_test_hosts=target.allow_test_hosts,
            )
        if (
            isinstance(request_interval_seconds, bool)
            or not isinstance(request_interval_seconds, (int, float))
            or not float(request_interval_seconds) > 0
        ):
            raise ValueError(
                "ManagerBasketAdapter: request_interval_seconds must be greater than zero: "
                f"{request_interval_seconds!r}"
            )
        if isinstance(retry_count, bool) or not isinstance(retry_count, int) or retry_count < 0:
            raise ValueError(
                f"ManagerBasketAdapter: retry_count must be a nonnegative integer: {retry_count!r}"
            )
        self.request_interval_seconds = float(request_interval_seconds)
        self.retry_count = retry_count
        self._next_request_at = 0.0
        self.window_start = window_start
        self.window_end = window_end
        self.cutoff = cutoff

    def discover(
        self,
        start: date | None = None,
        end: date | None = None,
        cutoff: datetime | None = None,
    ) -> list[ManagerBasketTarget]:
        """Targets with as_of inside [start, end] and published_at <= cutoff.

        Arguments fall back to the window configured on the adapter; a bound
        that is None everywhere is not applied. The result is sorted by
        (as_of, manager_code, fund_code) for deterministic batching.
        """
        effective_start = start if start is not None else self.window_start
        effective_end = end if end is not None else self.window_end
        effective_cutoff = cutoff if cutoff is not None else self.cutoff
        selected = [
            target
            for target in self._targets
            if (effective_start is None or effective_start <= target.as_of)
            and (effective_end is None or target.as_of <= effective_end)
            and (effective_cutoff is None or target.published_at <= effective_cutoff)
        ]
        return sorted(selected, key=lambda target: (target.as_of, target.manager_code, target.fund_code))

    def fetch(self, target: ManagerBasketTarget) -> Path:
        """GET target.source_url and store it atomically at the deterministic raw path.

        An existing nonempty raw file is returned without any network call.
        Requests are spaced at least request_interval_seconds apart and retried
        on 429/500/502/503/504 with bounded exponential backoff.
        """
        if not self.refresh:
            raise ValueError("ManagerBasketAdapter: explicit refresh is required for network collection")
        accept = _CSV_ACCEPT if target.format_hint == "csv" else _XLSX_ACCEPT
        for attempt in range(self.retry_count + 1):
            self._respect_rate_limit()
            try:
                payload = read_with_validated_redirects(
                    target.source_url,
                    headers={"User-Agent": self.user_agent, "Accept": accept},
                    timeout=_REQUEST_TIMEOUT_SECONDS,
                    validate_url=lambda url: validate_source_url(
                        MANAGER_BASKET_SOURCE,
                        url,
                        manager_hosts=target.reviewed_hosts,
                        allow_test_hosts=target.allow_test_hosts,
                    ),
                )
                digest = hashlib.sha256(payload).hexdigest()
                raw_path = self._raw_path(target, digest)
                if not raw_path.exists():
                    _atomic_write(raw_path, payload)
                return raw_path
            except urllib.error.HTTPError as error:
                if error.code in _RETRIABLE_STATUS_CODES and attempt < self.retry_count:
                    backoff = min(_BACKOFF_BASE_SECONDS * (2**attempt), _BACKOFF_MAX_SECONDS)
                    time.sleep(backoff)
                    continue
                raise
        raise AssertionError("unreachable: retry loop always returns or raises")

    def normalize(
        self,
        target: ManagerBasketTarget,
        raw_path: Path,
        output_dir: Path,
    ) -> tuple[Path, Path, int, int]:
        """Normalize one raw basket document into holdings JSONL plus quarantine JSONL.

        Returns (records_path, quarantine_path, records_count,
        quarantine_count). Both files are deterministically overwritten under
        output_dir/manager_basket/. The document is rejected with ValueError
        when any published per-row as-of date does not match target.as_of.
        """
        if target.format_hint not in FORMAT_HINTS:
            raise ValueError(
                f"manager basket document for {self._document_id(target)} has unsupported "
                f"format_hint {target.format_hint!r}; supported: {list(FORMAT_HINTS)}"
            )
        rows = _read_rows(target, Path(raw_path))
        for row in rows:
            if row.as_of_raw is None:
                continue
            document_label = self._document_id(target)
            row_as_of = _parse_as_of(row.as_of_raw)
            if row_as_of != target.as_of:
                raise ValueError(
                    f"manager basket document {document_label} reports holdings as of "
                    f"{row_as_of.isoformat()} (row {row.row_number}), which does not match "
                    f"the target as_of {target.as_of.isoformat()}; rejecting the document"
                )

        total_market_value = _document_total_market_value(rows)

        raw_sha256 = file_sha256(Path(raw_path))
        source_document_id = self._document_id(target, raw_sha256)
        records: list[HoldingsRecord] = []
        quarantined: list[BasketQuarantineEntry] = []
        for row in rows:
            outcome = self._normalize_row(target, row, total_market_value, source_document_id)
            if isinstance(outcome, BasketQuarantineEntry):
                quarantined.append(outcome)
            else:
                records.append(outcome)

        artifact_dir = (
            Path(output_dir)
            / MANAGER_BASKET_SOURCE
            / target.manager_code
            / target.fund_code
            / target.as_of.isoformat()
        )
        records_payload, record_count = serialize_jsonl(records)
        quarantine_payload, quarantine_count = _serialize_quarantine_jsonl(quarantined)
        records_path = publish_immutable(artifact_dir / "records", records_payload, ".jsonl")
        quarantine_path = publish_immutable(
            artifact_dir / "quarantine", quarantine_payload, ".quarantine.jsonl"
        )
        return records_path, quarantine_path, record_count, quarantine_count

    def _normalize_row(
        self,
        target: ManagerBasketTarget,
        row: _Row,
        total_market_value: float | None,
        source_document_id: str,
    ) -> HoldingsRecord | BasketQuarantineEntry:
        """Resolve, validate, and weigh one basket row; quarantine instead of crashing."""
        resolution = self._resolver.resolve(
            source_isin=row.source_isin,
            local_key_type=_KRX_CODE if row.krx_code else None,
            local_key=row.krx_code,
        )
        if resolution.isin is None:
            return self._quarantine(target, row, source_document_id, f"constituent identity unresolved: {resolution.reason}")
        if not row.name:
            return self._quarantine(target, row, source_document_id, "constituent name is missing")
        for label, raw, value in (
            ("weight percent", row.weight_pct_raw, row.weight_pct),
            ("quantity", row.quantity_raw, row.quantity),
            ("market value", row.market_value_raw, row.market_value),
        ):
            if raw is not None and value is None:
                return self._quarantine(target, row, source_document_id, f"{label} is not a valid number: {raw!r}")
        if row.quantity is not None and row.quantity < 0:
            return self._quarantine(target, row, source_document_id, f"quantity is negative: {row.quantity}")
        if row.market_value is not None and row.market_value < 0:
            return self._quarantine(target, row, source_document_id, f"market value is negative: {row.market_value}")

        if row.weight_pct is not None:
            weight = row.weight_pct / 100.0
            weight_source = _SOURCE_PUBLISHED
            if weight < 0:
                return self._quarantine(target, row, source_document_id, f"weight {weight} from weight percent {row.weight_pct} is negative")
            if weight > 1:
                return self._quarantine(target, row, source_document_id, f"weight {weight} from weight percent {row.weight_pct} exceeds 1.0")
        else:
            if row.market_value is None:
                return self._quarantine(target, row, source_document_id, "weight percent and market value are both missing; weight cannot be determined")
            if total_market_value is None:
                return self._quarantine(target, row, source_document_id, "weight percent is missing and no positive market values exist in this "
                "document to derive a document total from")
            weight = row.market_value / total_market_value
            weight_source = _DERIVED_FROM_VALUE

        if row.currency is not None:
            source_currency = row.currency
        elif row.market_value is not None:
            source_currency = _DEFAULT_CURRENCY
        else:
            source_currency = None

        record = HoldingsRecord.create(
            fund_isin=target.fund_isin,
            constituent_isin=resolution.isin,
            constituent_name=row.name,
            weight=weight,
            as_of=target.as_of,
            weight_source=weight_source,
            identifier_method=resolution.method,
            source_document_id=source_document_id,
            source_url=target.source_url,
            evidence_basis="manager_published",
            source_row_id=f"row:{row.row_number}",
            source_quantity=row.quantity,
            source_currency=source_currency,
            source_market_value=row.market_value,
            published_at=target.published_at,
        )
        return record

    def _quarantine(
        self,
        target: ManagerBasketTarget,
        row: _Row,
        source_document_id: str,
        reason: str,
    ) -> BasketQuarantineEntry:
        identifier_type = "isin" if row.source_isin else (_KRX_CODE if row.krx_code else None)
        identifier_value = row.source_isin if row.source_isin else row.krx_code
        return BasketQuarantineEntry(
            source_document_id=source_document_id,
            constituent_name=row.name,
            reason=f"row {row.row_number}: {reason}",
            source_identifier_type=identifier_type,
            source_identifier_value=identifier_value,
        )

    def _document_id(self, target: ManagerBasketTarget, raw_sha256: str | None = None) -> str:
        stable = f"{target.manager_code}:{target.fund_code}:{target.as_of.isoformat()}"
        return f"{stable}:{raw_sha256}" if raw_sha256 is not None else stable

    def _raw_path(self, target: ManagerBasketTarget, raw_sha256: str) -> Path:
        return (
            self.raw_root
            / MANAGER_BASKET_SOURCE
            / target.manager_code
            / target.fund_code
            / target.as_of.isoformat()
            / f"{raw_sha256}.{target.format_hint}"
        )

    def _respect_rate_limit(self) -> None:
        """Sleep so that requests are spaced at least request_interval_seconds apart."""
        now = time.monotonic()
        wait = self._next_request_at - now
        if wait > 0:
            time.sleep(wait)
            now = self._next_request_at
        self._next_request_at = max(now, self._next_request_at) + self.request_interval_seconds
